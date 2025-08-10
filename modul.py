import pandas as pd
import numpy as np
import math
import bisect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Union
from collections import defaultdict

class DataFilterAndSelect:
    COLUMNS = ["Kode Testpit", "Grid", "Prospek", "Tanggal Sampling", "Total Kedalaman",
               "Total Koli", "Pemilik Lahan", "Penggali", "Pengangkut", "Penimbun"]
           
    def __init__(self, source):
        if isinstance(source, pd.DataFrame):
            self.df = source
        elif isinstance(source, str):  # assume filename
            try:
                self.df = pd.read_csv(source, encoding='utf-8')
            except Exception as e:
                raise ValueError(f"Failed to load CSV file '{source}': {e}")
        else:
            raise ValueError("Unsupported input type for DataFilterAndSelect")

        # Konversi kolom tanggal jika ada
        if "Tanggal Sampling" in self.df.columns:
            self.df["Tanggal Sampling"] = pd.to_datetime(
                self.df["Tanggal Sampling"], errors='coerce', dayfirst=True
            )

        self.cleanData = None
        
    def filter_and_select(self):
        filtered_df = self.df.loc[self.df["Total Kedalaman"].notna(), self.COLUMNS].copy()
        if filtered_df.empty:
            raise ValueError("Kolom Contoh Error")
        
        self.cleanData = filtered_df
        return self.cleanData


class ConfigurationInput:
    def __init__(self):
        self.stage1 = None
        self.stage2 = None
        self.stage3 = None

    @staticmethod
    def _merge_stage_data(existing_df, new_df, subset):
        if existing_df is None:
            return new_df
        return pd.concat([existing_df, new_df], ignore_index=True).drop_duplicates(subset=subset)

    def process_stage1(self, cleanData):
        unique_locations = cleanData["Prospek"].unique()
        new_data = pd.DataFrame({
            "Lokasi": unique_locations,
            "Tanggal Mulai (2025-05-23)": pd.NaT,
            "Tanggal Selesai (2025-05-23)": pd.NaT,
            "Tanggal Gajian (2025-05-23)": pd.NaT,
            "Sistem Angkutan (Koli/Kilo)": np.nan,
        })

        # Explicitly convert those date columns to datetime dtype (with normalized time)
        date_cols = [
            "Tanggal Mulai (2025-05-23)",
            "Tanggal Selesai (2025-05-23)",
            "Tanggal Gajian (2025-05-23)",
        ]

        self.stage1 = self._merge_stage_data(self.stage1, new_data, subset=["Lokasi"])
        return self.stage1

    def _filter_by_location_and_date(self, cleanData, stage1_data):
        filtered_chunks = []

        for _, row in stage1_data.iterrows():
            lokasi = row["Lokasi"]
            tgl_mulai = row["Tanggal Mulai (2025-05-23)"]
            tgl_selesai = row["Tanggal Selesai (2025-05-23)"]

            print(f"🔍 Filtering for Lokasi: {lokasi} | Start: {tgl_mulai} | End: {tgl_selesai}")

            if pd.isna(tgl_mulai) and pd.isna(tgl_selesai):
                continue  # skip rows without a date filter

            filtered = cleanData[cleanData["Prospek"] == lokasi]

            if pd.notna(tgl_mulai) and pd.notna(tgl_selesai):
                filtered = filtered[(filtered["Tanggal Sampling"] >= tgl_mulai) & (filtered["Tanggal Sampling"] <= tgl_selesai)]
            elif pd.notna(tgl_mulai):
                filtered = filtered[filtered["Tanggal Sampling"] >= tgl_mulai]
            elif pd.notna(tgl_selesai):
                filtered = filtered[filtered["Tanggal Sampling"] <= tgl_selesai]

            print(f"✅ Matched rows for '{lokasi}': {len(filtered)}")

            if not filtered.empty:
                filtered_chunks.append(filtered)

        if filtered_chunks:
            result = pd.concat(filtered_chunks, ignore_index=True)
            print(f"🔍 Filtered cleanData rows: {len(result)}")
            return result
        else:
            print("⚠️ No data matched any filters.")
            return pd.DataFrame(columns=cleanData.columns)

    def process_stage2(self, cleanData, stage1_data):
        filtered_data = self._filter_by_location_and_date(cleanData, stage1_data)
        if filtered_data.empty:
            print("⚠️ Stage 2: No matching rows found after filter.")
            return pd.DataFrame(columns=["Penggali", "Kelompok Penggali", "Harga Galian (Lokal/Luar)", "Harga Samplingan (Lokal/Luar)"])

        penggali_gajian = filtered_data["Penggali"].unique()
        new_data = pd.DataFrame({
            "Penggali": penggali_gajian,
            "Kelompok Penggali": np.nan,
            "Harga Galian (Lokal/Luar)": np.nan,
            "Harga Samplingan (Lokal/Luar)": np.nan,
        })
        self.stage2 = self._merge_stage_data(self.stage2, new_data, subset=["Penggali"])
        return self.stage2

    def process_stage3(self, cleanData, stage1_data):
        result = self._filter_by_location_and_date(cleanData, stage1_data)
    
        # Create mapping from Lokasi to Sistem Angkutan
        mapping = stage1_data.set_index("Lokasi")["Sistem Angkutan (Koli/Kilo)"].to_dict()
    
        # Map to result DataFrame
        result["SistemAngkutan"] = result["Prospek"].map(mapping)
    
        self.stage3 = result
        return result

class PaymentCount:
    def __init__(self, harga_galian_lokal, harga_galian_luar,
                 harga_samplingan_lokal, harga_samplingan_luar):
        self.harga_galian_lokal = harga_galian_lokal
        self.harga_galian_luar = harga_galian_luar
        self.harga_samplingan_lokal = harga_samplingan_lokal
        self.harga_samplingan_luar = harga_samplingan_luar
        self.df = None

    def set_data(self, df):
        self.df = df.copy()
        return self

    def harga_galian(self):
        def merge_row(row):
            key = "Harga Galian (Lokal/Luar)"
            if key not in row or pd.isna(row[key]):
                return row

            if row[key].strip().lower() == "luar":
                harga_row = self.harga_galian_luar[self.harga_galian_luar["Kedalaman"] == row["Total Kedalaman"]]
            else:
                harga_row = self.harga_galian_lokal[self.harga_galian_lokal["Kedalaman"] == row["Total Kedalaman"]]

            if not harga_row.empty:
                row["Tarif Galian"] = harga_row.iloc[0]["Harga"]
            return row

        self.df = self.df.apply(merge_row, axis=1)
        return self

    def harga_samplingan(self):
        def merge_row(row):
            key = "Harga Samplingan (Lokal/Luar)"
            
            # NEW: Default to 0 if key is missing or NaN
            if key not in row or pd.isna(row[key]):
                row["Tarif Samplingan"] = 0
                return row
    
            if row[key].strip().lower() == "luar":
                harga_row = self.harga_samplingan_luar[self.harga_samplingan_luar["Total Koli"] == row["Total Koli"]]
            else:
                harga_row = self.harga_samplingan_lokal[self.harga_samplingan_lokal["Total Koli"] == row["Total Koli"]]
    
            if not harga_row.empty:
                row["Tarif Samplingan"] = harga_row.iloc[0]["Harga"]
            else:
                # NEW: Set to 0 when no match is found
                row["Tarif Samplingan"] = 0
    
            return row
    
        self.df = self.df.apply(merge_row, axis=1)
        return self

    def harga_timbunan_dan_kompensasi_langsiran(self):
        self.df["Tarif Timbunan"] = self.df["Total Kedalaman"] * 12000
        self.df["Tarif Kompensasi"] = 90000
        self.df["Tarif Langsiran"] = self.df["Penimbun"] * 1000 * self.df["Total Koli"]
        return self

    def harga_angkutan(self):
        def hitung_angkutan(row):
            sistem = str(row.get("SistemAngkutan", "")).strip().lower()
            
            if sistem == "koli":
                return row.get("Total Koli", 0) * row.get("Pengangkut", 0) * 1000  
            elif sistem == "kilo":
                return row.get("Pengangkut", 0) * 1000   
            else:
                return 0  
    
        self.df["Tarif Angkutan"] = self.df.apply(hitung_angkutan, axis=1)
        return self

    def get_result(self):
        if "Kode Testpit" in self.df.columns:
            self.df = self.df.drop_duplicates(subset=["Kode Testpit"])
        return self.df 
    
    def get_pivot_summary(self):
        # Create pivot table
        pivot_df = self.df.pivot_table(
            index=['Tanggal Sampling', 'Kode Testpit', 'Grid', 'Prospek', 'Penggali', 'Pemilik Lahan'],
            values=['Tarif Galian', 'Tarif Samplingan', 'Tarif Timbunan',
                    'Tarif Kompensasi', 'Tarif Angkutan', 'Tarif Langsiran'],
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
            
        # Select numeric columns as a list
        numeric_cols = list(pivot_df.select_dtypes(include='number').columns)
    
        # Add Total column (row-wise sum)
        pivot_df['Total'] = pivot_df[numeric_cols].sum(axis=1)
    
        # Create total row
        total_values = pivot_df[numeric_cols + ['Total']].sum(axis=0)
        total_row = {col: total_values.get(col, None) for col in pivot_df.columns}
    
        # Fill non-numeric columns with "Total"
        for col in pivot_df.columns:
            if col not in numeric_cols and col != 'Total':
                total_row[col] = 'Total'
                
        for col in pivot_df.columns:
            if col not in numeric_cols + ['Total']:
                pivot_df[col] = pivot_df[col].astype(str)
    
        # Append as a new row (numeric index stays clean)
        pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])], ignore_index=True)
    
        return pivot_df

class MultiPaymentExcel:
    def __init__(
        self,
        ws,
        data_rows: List[List[List[Union[str, float, int]]]],
        group_names: List[str],
        date_text: str = "Setabar, 26 Juni 2025",
        signers: dict = None ,
        receiver_title: str = "Area",
        mode: str = "gali"
    ):
        if signers is None:
            signers = {
                "B": ("Chandra Ardiansyah", "Keu. / Umum"),
                "D": ("Rizky Lambas", "Geologist"),
            }        
        self.ws = ws
        self.data_rows = data_rows
        self.group_names = group_names
        self.signers = signers
        self.date_text = date_text
        self.receiver_title = receiver_title
        self.mode = mode.lower()

    def generate_excel(self):
        ws = self.ws
        current_row = 1

        # Styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        rp_style = NamedStyle(name="rupiah_style")
        rp_style.number_format = '"Rp."#,##0'
        rp_style.alignment = center_align
        rp_style.border = thin_border
        if "rupiah_style" not in ws.parent.named_styles:
            ws.parent.add_named_style(rp_style)

        # Mode-specific config
        mode_config = {
            "gali": {
                "title": "PENGGALIAN TEST PIT",
                "uraian": "Untuk Pembayaran Penggalian Test Pit sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Kedalaman (m)", "Harga Borongan"],
                "harga_col": 6,
                "subtotal": False,
            },
            "sampling": {
                "title": "PENYAMPLINGAN TEST PIT",
                "uraian": "Untuk Pembayaran Penyamplingan Test Pit sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Total Koli", "Harga Borongan"],
                "harga_col": 6,
                "subtotal": False,
            },
            "timbunan": {
                "title": "PEMBAYARAN TIMBUNAN TEST PIT",
                "uraian": "Untuk Pembayaran Timbunan Test Pit sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Kedalaman (m)", "Harga Borongan", "TTD"],
                "harga_col": 8,
                "subtotal": False,
            },
            "kompensasi": {
                "title": "PEMBAYARAN KOMPENSASI LAHAN",
                "uraian": "Untuk Pembayaran Kompensasi Lahan sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Harga Kompensasi", "Total Kompensasi", "TTD"],
                "harga_col": 7,
                "subtotal": True,
            },
            "angkutan": {
                "title": "PEMBAYARAN ANGKUTAN SAMPEL",
                "uraian": "Untuk Pembayaran Angkutan Sampel sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Harga Angkutan", "TTD"],
                "harga_col": 7,
                "subtotal": False,
            },
            "langsiran": {
                "title": "PEMBAYARAN LANGSIRAN SAMPEL",
                "uraian": "Untuk Pembayaran Langsiran Sampel sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Harga Langsiran", "TTD"],
                "harga_col": 7,
                "subtotal": False,
            }
            
        }

        config = mode_config[self.mode]

        for table_index, table_rows in enumerate(self.data_rows):
            group_name = self.group_names[table_index]

            for title in ["BUKTI PEMBAYARAN", config["title"]]:
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
                cell = ws.cell(row=current_row, column=2, value=title)
                cell.font = Font(bold=True, size=14 if title == "BUKTI PEMBAYARAN" else 12)
                cell.alignment = center_align
                current_row += 1

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
            ws.cell(row=current_row, column=2, value="Sudah Terima Dari : Tim Eksplorasi Bauksit Kalbar").alignment = left_align
            current_row += 2

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
            ws.cell(row=current_row, column=2, value=config["uraian"]).alignment = left_align
            current_row += 1

            for col_index, header in enumerate(config["headers"], start=2):
                cell = ws.cell(row=current_row, column=col_index, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
                ws.column_dimensions[get_column_letter(col_index)].width = len(header) + 5
            current_row += 1

            total_harga = 0
            start_data_row = current_row
            # Sort table_rows based on mode category
            if self.mode in ["gali", "sampling"]:
                table_rows
            elif self.mode in ["timbunan", "angkutan", "kompensasi"]:
                # Sort by pemilik lahan (index 4)
                table_rows.sort(key=lambda x: str(x[4]).strip().lower())

            for i, row_data in enumerate(table_rows):
                row_data[0] = i + 1
                while len(row_data) < len(config["headers"]):
                    row_data.append("")
                for j, val in enumerate(row_data):
                    col = j + 2
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.alignment = center_align
                    cell.border = thin_border
                    if col == config["harga_col"]:
                        cell.style = "rupiah_style"
                if isinstance(row_data[config["harga_col"] - 2], (int, float)):
                    total_harga += row_data[config["harga_col"] - 2]
                current_row += 1

            # Subtotal for kompensasi
            if self.mode == "kompensasi":
                owner_row_map = defaultdict(list)
                for idx, row_data in enumerate(table_rows):
                    owner = str(row_data[4]).strip()
                    owner_row_map[owner].append(start_data_row + idx)
                for owner, rows in owner_row_map.items():
                    subtotal = 0
                    for row in rows:
                        cell = ws.cell(row=row, column=7)
                        if isinstance(cell.value, (int, float)):
                            subtotal += cell.value
                    ws.cell(row=rows[0], column=8, value=subtotal).style = "rupiah_style"

            # Grand total row
            total_label_col = config["harga_col"] - 1
            label_cell = ws.cell(row=current_row, column=total_label_col, value="TOTAL")
            label_cell.font = bold_font
            label_cell.alignment = center_align
            label_cell.border = thin_border

            total_cell = ws.cell(row=current_row, column=config["harga_col"], value=total_harga)
            total_cell.style = "rupiah_style"

            current_row += 2

            if self.mode in ["gali", "sampling"]:
                ws.cell(row=current_row, column=5, value=self.date_text).alignment = left_align
                current_row += 1
                ws.cell(row=current_row, column=2, value="Dibayar Oleh,").alignment = center_align
                ws.cell(row=current_row, column=4, value="Pet. Lapangan,").alignment = center_align
                ws.cell(row=current_row, column=6, value="Yang Menerima,").alignment = center_align
                current_row += 5
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][0]).alignment = center_align
                ws.cell(row=current_row, column=4, value=self.signers["D"][0]).alignment = center_align
                ws.cell(row=current_row, column=6, value=group_name).alignment = center_align
                current_row += 1
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][1]).alignment = center_align
                ws.cell(row=current_row, column=4, value=self.signers["D"][1]).alignment = center_align
                ws.cell(row=current_row, column=6, value="Ketua Kelompok").alignment = center_align
                current_row += 4

            if self.mode in ["kompensasi", "timbunan", "angkutan"]:
                ws.cell(row=current_row, column=7, value=self.date_text).alignment = left_align
                current_row += 1
                ws.cell(row=current_row, column=2, value="Dibayar Oleh,").alignment = center_align
                ws.cell(row=current_row, column=5, value="Pet. Lapangan,").alignment = center_align
                ws.cell(row=current_row, column=8, value="Lokasi,").alignment = center_align
                current_row += 5
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][0]).alignment = center_align
                ws.cell(row=current_row, column=5, value=self.signers["D"][0]).alignment = center_align
                ws.cell(row=current_row, column=8, value=group_name).alignment = center_align
                current_row += 1
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][1]).alignment = center_align
                ws.cell(row=current_row, column=5, value=self.signers["D"][1]).alignment = center_align
                ws.cell(row=current_row, column=8, value=self.receiver_title).alignment = center_align
                current_row += 4
                
class PaymentExcelBuilder:
    def __init__(self, df: pd.DataFrame):
        self.df = df.sort_values(by=["Kelompok Penggali", 'Penggali']).copy()
        self.df.fillna(0, inplace=True)
        
    def _group_data(self, group_col, columns, rename_map, values_structure, mode: str):
        raw_data = self.df[columns].rename(columns=rename_map)
        raw_data_list = raw_data.to_dict(orient="records")
    
        grouped = defaultdict(list)
        for entry in raw_data_list:
            row = [
                None,
                *[entry[col] for col in values_structure]
            ]
    
            # Add signature column only for kompensasi or timbunan
            if mode in ("kompensasi", "timbunan", "angkutan"):
                row.append("")
    
            grouped[entry[group_col]].append(row)
    
        tables = list(grouped.values())
        names = list(grouped.keys())
        return tables, names
    def create_multi_payment_excel(
        self,
        output_file: str,
        date_text: str = "Setabar, 26 Juni 2025",
        signers: dict = None
    ):
        if signers is None:
            signers = {
                "B": ("Chandra Ardiansyah", "Keu. / Umum"),
                "D": ("Rizky Lambas", "Geologist"),
            }
        wb = Workbook()
        wb.remove(wb.active)

        configs = [
            {
                "sheet": "Galian",
                "mode": "gali",
                "group_col": "Penggali",
                "columns": ["Penggali", "Tanggal Sampling", "Kode Testpit", "Total Kedalaman", "Tarif Galian"],
                "rename": {"Tarif Galian": "tarif"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Total Kedalaman", "tarif"]
            },
            {
                "sheet": "Samplingan",
                "mode": "sampling",
                "group_col": "Penggali",
                "columns": ["Penggali", "Tanggal Sampling", "Kode Testpit", "Total Koli", "Tarif Samplingan"],
                "rename": {"Tarif Samplingan": "tarif"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Total Koli", "tarif"]
            },
            {
                "sheet": "Timbunan",
                "mode": "timbunan",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Total Kedalaman", "Tarif Timbunan"],
                "rename": {"Tarif Timbunan": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Total Kedalaman", "harga"]
            },
            {
                "sheet": "Kompensasi",
                "mode": "kompensasi",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Tarif Kompensasi"],
                "rename": {"Tarif Kompensasi": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "harga"]
            },
            {
                "sheet": "Angkutan",
                "mode": "angkutan",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Tarif Angkutan"],
                "rename": {"Tarif Angkutan": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "harga"]
            },
            {   "sheet": "Langsiran",
                "mode": "langsiran",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Tarif Langsiran"],
                "rename": {"Tarif Langsiran": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "harga"]
            }
            
        ]

        for config in configs:
            tables, names = self._group_data(
                config["group_col"],
                config["columns"],
                config["rename"],
                config["values"],
                config["mode"]
            )
            ws = wb.create_sheet(config["sheet"])
            report = MultiPaymentExcel(
                ws,
                tables,
                names,
                date_text=date_text,
                signers=signers,
                mode=config["mode"]
            )
            report.generate_excel()

        wb.save(output_file)
